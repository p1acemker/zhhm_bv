from datetime import datetime
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from tools.build_description_assets import (
    REQUIRED_DICTIONARY_SHEETS,
    build_assets,
    derive_mature_rules,
)


HEADERS = [
    "生产公司",
    "客户简称",
    "单证客户名称",
    "合同日期",
    "订单号",
    "订单类型",
    "销售部门",
    "产品编码",
    "客户货号",
    "品种",
    "规格",
    "材质分类",
    "表面处理",
    "英文描述",
    "工厂交货期",
]


def _write_source(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "sheet1"
    sheet.append(HEADERS)
    sheet.append(
        [
            "A",
            "ACME",
            "ACME LTD",
            datetime(2024, 8, 1),
            "O-1",
            "NORMAL",
            "EXPORT",
            "RXD381X_D100_91VQ11R40",
            "",
            "RXD381X",
            "D100",
            "Q11",
            "R40",
            'DI LUG WAFER BV, EPDM SEAT, SS316 DISC, GEAR, PN16, 4"/DN100',
            datetime(2024, 9, 1),
        ]
    )
    sheet.append(
        [
            "A",
            "ACME",
            "ACME LTD",
            datetime(2024, 8, 2),
            "O-2",
            "NORMAL",
            "EXPORT",
            "30199010115",
            "",
            "MXHQ",
            "D50",
            "L02",
            "",
            'SUPERVISORY SWITCH FOR OS&Y GATE VALVE 2"-16"',
            datetime(2024, 9, 2),
        ]
    )
    sheet.append(
        [
            "A",
            "ACME",
            "ACME LTD",
            datetime(2024, 8, 3),
            "O-3",
            "NORMAL",
            "EXPORT",
            "Q11_D50_900Q11R40",
            "",
            "Q11",
            "D50",
            "Q11",
            "R40",
            'BRONZE BALL VALVE 2"',
            datetime(2024, 9, 3),
        ]
    )
    workbook.save(path)


def test_derive_mature_rules_requires_stable_train_and_validation() -> None:
    records = []
    for index in range(20):
        records.append(
            {
                "date": datetime(2023, 1, 1),
                "by1": "D71X",
                "form_code": "90F",
                "attributes": {"connection": {"value": "WAFER"}},
            }
        )
    for index in range(5):
        records.append(
            {
                "date": datetime(2024, 3, 1),
                "by1": "D71X",
                "form_code": "90F",
                "attributes": {"connection": {"value": "WAFER"}},
            }
        )

    rules = derive_mature_rules(records)

    assert rules[("D71X", "90F")]["connection"]["value"] == "WAFER"
    assert rules[("D71X", "90F")]["connection"]["confidence"] == 0.99


def test_build_assets_preserves_source_rows_and_creates_reviewable_outputs(
    tmp_path: Path,
) -> None:
    source = tmp_path / "orders.xlsx"
    output = tmp_path / "output"
    _write_source(source)

    result = build_assets(source, output, dictionary_version="test-v1")

    standardized = load_workbook(result["standardized_workbook"], read_only=True)
    sheet = standardized.active
    rows = list(sheet.iter_rows(values_only=True))
    assert len(rows) == 4
    assert list(rows[0][: len(HEADERS)]) == HEADERS
    source_rows = list(
        load_workbook(source, read_only=True).active.iter_rows(values_only=True)
    )
    assert list(rows[1][: len(HEADERS)]) == list(source_rows[1])
    output_headers = list(rows[0])
    description_index = output_headers.index("standardized_description")
    role_index = output_headers.index("product_role")
    assert "BUTTERFLY VALVE" in rows[1][description_index]
    assert rows[2][role_index] == "accessory"
    assert rows[2][description_index] is None
    assert rows[3][role_index] == "other"

    dictionary = load_workbook(result["dictionary_workbook"], read_only=True)
    assert set(REQUIRED_DICTIONARY_SHEETS).issubset(dictionary.sheetnames)

    compiled = json.loads(Path(result["dictionary_json"]).read_text(encoding="utf-8"))
    assert compiled["version"] == "test-v1"
    assert len(compiled["source_sha256"]) == 64

    gold = load_workbook(result["gold_workbook"], read_only=True)
    gold_headers = [cell.value for cell in next(gold.active.iter_rows())]
    assert "approved" in gold_headers
    assert "corrected_description" in gold_headers
    assert "corrected_body_material" in gold_headers
    assert "corrected_template_id" in gold_headers


def test_build_assets_preserves_business_managed_dictionary_terms(tmp_path: Path) -> None:
    source = tmp_path / "orders.xlsx"
    first_output = tmp_path / "first"
    second_output = tmp_path / "second"
    _write_source(source)
    first = build_assets(source, first_output, dictionary_version="v1")
    dictionary = load_workbook(first["dictionary_workbook"])
    dictionary["词语映射"].append(
        ["CUSTOM BODY", "DUCTILE IRON", "body_material", "butterfly", 99, True, "business maintained"]
    )
    dictionary.save(first["dictionary_workbook"])

    second = build_assets(
        source,
        second_output,
        dictionary_version="v2",
        dictionary_source=first["dictionary_workbook"],
    )

    rebuilt = load_workbook(second["dictionary_workbook"], read_only=True, data_only=True)
    rows = list(rebuilt["词语映射"].iter_rows(values_only=True))
    assert any(row[0] == "CUSTOM BODY" for row in rows)
