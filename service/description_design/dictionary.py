"""Business vocabulary used by the valve description engine."""

from __future__ import annotations

from dataclasses import dataclass
import json
from pathlib import Path
import re
from typing import Iterable


@dataclass(frozen=True)
class TermRule:
    """Map one source phrase to a canonical business-field value."""

    term: str
    value: str
    field: str
    families: tuple[str, ...] = ()
    priority: int = 0


@dataclass(frozen=True)
class TermMatch:
    """Winning dictionary match for one business field."""

    value: str
    source_term: str
    priority: int


@dataclass(frozen=True)
class ExceptionRule:
    """Override one field when an approved business condition matches."""

    match_field: str
    match_value: str
    override_field: str
    override_value: str
    reason: str = ""


def _phrase_pattern(term: str) -> re.Pattern[str]:
    escaped = re.escape(term.upper()).replace(r"\ ", r"\s+")
    return re.compile(rf"(?<![A-Z0-9]){escaped}(?![A-Z0-9])")


class BusinessDictionary:
    """Apply boundary-aware, family-scoped, longest-phrase matching."""

    def __init__(
        self,
        rules: Iterable[TermRule],
        exceptions: Iterable[ExceptionRule] = (),
    ) -> None:
        self.rules = tuple(rules)
        self.exceptions = tuple(exceptions)

    @classmethod
    def default(cls) -> "BusinessDictionary":
        """Return the built-in bootstrap dictionary."""
        rules = [
            TermRule("DUCTILE IRON", "DUCTILE IRON", "body_material", priority=20),
            TermRule("DI", "DUCTILE IRON", "body_material", priority=10),
            TermRule("CAST IRON", "CAST IRON", "body_material", priority=20),
            TermRule("CI", "CAST IRON", "body_material", priority=10),
            TermRule("CARBON STEEL", "CARBON STEEL", "body_material", priority=20),
            TermRule("CS", "CARBON STEEL", "body_material", priority=10),
            TermRule("LUG WAFER", "LUG WAFER", "connection", priority=30),
            TermRule("LUGGED WAFER", "LUG WAFER", "connection", priority=30),
            TermRule("GROOVED", "GROOVED", "connection", priority=20),
            TermRule("GRVD", "GROOVED", "connection", priority=20),
            TermRule("GRV", "GROOVED", "connection", priority=10),
            TermRule("FLANGED", "FLANGED", "connection", priority=20),
            TermRule("FLGD", "FLANGED", "connection", priority=20),
            TermRule("FLG", "FLANGED", "connection", priority=10),
            TermRule("THREADED", "THREADED", "connection", priority=20),
            TermRule("THD", "THREADED", "connection", priority=10),
            TermRule("WAFER", "WAFER", "connection", priority=10),
            TermRule("LUG", "LUG", "connection", priority=10),
            TermRule("EPDM SEAT", "EPDM", "seat_material", priority=20),
            TermRule("EPDM", "EPDM", "seat_material", priority=10),
            TermRule("BUNA-N", "NBR", "seat_material", priority=20),
            TermRule("NITRILE", "NBR", "seat_material", priority=20),
            TermRule("NBR", "NBR", "seat_material", priority=10),
            TermRule("PTFE", "PTFE", "seat_material", priority=10),
            TermRule("VITON", "VITON", "seat_material", priority=10),
            TermRule("DI+EPDM DISC", "DUCTILE IRON WITH EPDM COATED", "closure_material", priority=40),
            TermRule("DI + EPDM DISC", "DUCTILE IRON WITH EPDM COATED", "closure_material", priority=40),
            TermRule("SS316 DISC", "STAINLESS STEEL 316", "closure_material", priority=30),
            TermRule("316 SS DISC", "STAINLESS STEEL 316", "closure_material", priority=30),
            TermRule("SS304 DISC", "STAINLESS STEEL 304", "closure_material", priority=30),
            TermRule("304 SS DISC", "STAINLESS STEEL 304", "closure_material", priority=30),
            TermRule("DI DISC", "DUCTILE IRON", "closure_material", priority=20),
            TermRule("GEAR OPERATED", "GEAR OPERATED", "actuation", priority=30),
            TermRule("WORM GEAR", "GEAR OPERATED", "actuation", priority=30),
            TermRule("GEAR", "GEAR OPERATED", "actuation", priority=10),
            TermRule("HANDWHEEL", "HANDWHEEL OPERATED", "actuation", priority=20),
            TermRule("LEVER", "LEVER OPERATED", "actuation", priority=10),
            TermRule("MOTORIZED", "MOTORIZED", "actuation", priority=20),
            TermRule("PNEUMATIC", "PNEUMATIC OPERATED", "actuation", priority=20),
            TermRule("FIRE RISER", "FIRE RISER", "structure", ("butterfly",), 30),
            TermRule("TRIPLE ECCENTRIC", "TRIPLE ECCENTRIC", "structure", ("butterfly",), 30),
            TermRule("DOUBLE ECCENTRIC", "DOUBLE ECCENTRIC", "structure", ("butterfly",), 30),
            TermRule("DUAL PLATE", "DUAL PLATE", "structure", ("check",), 30),
            TermRule("D.D.", "DUAL PLATE", "structure", ("check",), 20),
            TermRule("SWING", "SWING", "structure", ("check",), 20),
            TermRule("SILENT", "SILENT", "structure", ("check",), 20),
            TermRule("BALL CHECK", "BALL", "structure", ("check",), 30),
            TermRule("OS&Y", "OS&Y", "structure", ("gate",), 30),
            TermRule("OSY", "OS&Y", "structure", ("gate",), 20),
            TermRule("NON-RISING STEM", "NON-RISING STEM", "structure", ("gate",), 30),
            TermRule("NRS", "NON-RISING STEM", "structure", ("gate",), 20),
            TermRule("KNIFE GATE", "KNIFE", "structure", ("gate",), 30),
            TermRule("AWWA C508", "AWWA C508", "standard_certification", ("check",), 30),
            TermRule("UL/FM", "UL LISTED / FM APPROVED", "standard_certification", priority=30),
            TermRule("UL FM", "UL LISTED / FM APPROVED", "standard_certification", priority=20),
        ]
        return cls(rules)

    @classmethod
    def from_compiled_json(cls, path: str | Path) -> "BusinessDictionary":
        """Load term rules compiled from the reviewable Excel dictionary."""
        payload = json.loads(Path(path).read_text(encoding="utf-8"))
        sheets = payload.get("sheets", {})
        rows = sheets.get("词语映射", [])
        rules = cls._term_rules_from_rows(rows)
        exceptions = cls._exception_rules_from_rows(sheets.get("例外规则", []))
        return cls(rules, exceptions)

    @classmethod
    def from_excel(cls, path: str | Path) -> "BusinessDictionary":
        """Load the business-maintained Excel source directly."""
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)

        def rows(sheet_name: str) -> list[dict[str, object]]:
            if sheet_name not in workbook.sheetnames:
                return []
            iterator = workbook[sheet_name].iter_rows(values_only=True)
            headers = [str(value or "").strip() for value in next(iterator)]
            return [
                {header: value for header, value in zip(headers, row)}
                for row in iterator
                if any(value is not None for value in row)
            ]

        rules = cls._term_rules_from_rows(rows("词语映射"))
        exceptions = cls._exception_rules_from_rows(rows("例外规则"))
        workbook.close()
        return cls(rules, exceptions)

    @staticmethod
    def _active(value: object) -> bool:
        return str(value).strip().lower() not in {"false", "0", "no", "off"}

    @classmethod
    def _term_rules_from_rows(cls, rows: Iterable[dict[str, object]]) -> list[TermRule]:
        rules = []
        for row in rows:
            if not cls._active(row.get("active", True)):
                continue
            term = str(row.get("raw_term") or "").strip()
            value = str(row.get("canonical_term") or "").strip()
            field = str(row.get("field") or "").strip()
            if not term or not value or not field:
                continue
            scope = tuple(
                item.strip().lower()
                for item in str(row.get("family_scope") or "").split(",")
                if item.strip()
            )
            rules.append(
                TermRule(
                    term=term,
                    value=value,
                    field=field,
                    families=scope,
                    priority=int(row.get("priority") or 0),
                )
            )
        return rules

    @classmethod
    def _exception_rules_from_rows(
        cls,
        rows: Iterable[dict[str, object]],
    ) -> list[ExceptionRule]:
        exceptions = []
        for row in rows:
            if not cls._active(row.get("active", True)):
                continue
            values = {
                key: str(row.get(key) or "").strip()
                for key in [
                    "match_field",
                    "match_value",
                    "override_field",
                    "override_value",
                    "reason",
                ]
            }
            if all(values[key] for key in ["match_field", "match_value", "override_field", "override_value"]):
                exceptions.append(ExceptionRule(**values))
        return exceptions

    def match_terms(
        self,
        text: str,
        family: str | None = None,
        context: dict[str, object] | None = None,
    ) -> dict[str, TermMatch]:
        """Return one winning match per field."""
        normalized = str(text).upper()
        candidates: dict[str, list[tuple[int, int, TermRule]]] = {}
        for rule in self.rules:
            if rule.families and family not in rule.families:
                continue
            if not _phrase_pattern(rule.term).search(normalized):
                continue
            candidates.setdefault(rule.field, []).append(
                (len(rule.term), rule.priority, rule)
            )
        matches: dict[str, TermMatch] = {}
        for field, values in candidates.items():
            _, _, rule = max(values, key=lambda item: (item[0], item[1]))
            matches[field] = TermMatch(rule.value, rule.term, rule.priority)
        resolved_context = {
            "description": normalized,
            "family": str(family or "").upper(),
            "valve_family": str(family or "").upper(),
            **{
                str(key).lower(): str(value or "").strip().upper()
                for key, value in (context or {}).items()
            },
        }
        for exception in self.exceptions:
            actual = resolved_context.get(exception.match_field.lower(), "")
            expected = exception.match_value.upper()
            matched = expected in actual if exception.match_field.lower() == "description" else actual == expected
            if matched:
                matches[exception.override_field] = TermMatch(
                    exception.override_value,
                    f"EXCEPTION:{exception.match_value}",
                    10000,
                )
        return matches
