"""Evaluate a completed valve-description gold-review workbook."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import load_workbook


CRITICAL_FIELDS = [
    "body_material",
    "seat_material",
    "closure_material",
    "pressure",
    "standard_certification",
    "size",
]


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _approved(value: object) -> bool | None:
    normalized = _text(value).lower()
    if normalized in {"yes", "y", "true", "1", "是", "通过"}:
        return True
    if normalized in {"no", "n", "false", "0", "否", "不通过"}:
        return False
    return None


def _truth(row: Mapping[str, Any], field: str, predicted_field: str | None = None) -> str:
    corrected = _text(row.get(f"corrected_{field}"))
    return corrected or _text(row.get(predicted_field or field))


def _ratio(correct: int, total: int) -> float:
    return round(correct / total, 4) if total else 0.0


def evaluate_review_rows(rows: Iterable[Mapping[str, Any]]) -> dict[str, Any]:
    """Calculate auditable metrics from reviewed row dictionaries."""
    reviewed = []
    for row in rows:
        approval = _approved(row.get("approved"))
        if approval is not None:
            reviewed.append((row, approval))

    family_correct = role_correct = description_approved = 0
    critical_correct = critical_total = 0
    autofill_correct = autofill_total = 0
    top1_correct = top3_correct = template_total = 0
    predicted_target = true_target_and_predicted = 0
    for row, approval in reviewed:
        predicted_family = _text(row.get("predicted_family"))
        true_family = _truth(row, "family", "predicted_family")
        predicted_role = _text(row.get("predicted_role"))
        true_role = _truth(row, "role", "predicted_role")
        family_correct += predicted_family == true_family
        role_correct += predicted_role == true_role
        description_approved += approval
        if predicted_role == "valve" and predicted_family in {"butterfly", "check", "gate"}:
            predicted_target += 1
            true_target_and_predicted += (
                true_role == "valve"
                and true_family in {"butterfly", "check", "gate"}
            )

        for field in CRITICAL_FIELDS:
            predicted = _text(row.get(field))
            truth = _truth(row, field)
            if predicted or truth:
                critical_total += 1
                critical_correct += predicted == truth
        inferred_fields = {
            item.strip()
            for item in _text(row.get("inferred_fields")).split(",")
            if item.strip()
        }
        for field in inferred_fields:
            autofill_total += 1
            autofill_correct += _text(row.get(field)) == _truth(row, field)

        predicted_template = _text(row.get("template_id"))
        true_template = _truth(row, "template_id", "template_id")
        if predicted_template or true_template:
            template_total += 1
            top1_correct += predicted_template == true_template
            alternatives = {
                item.strip()
                for item in _text(row.get("alternative_template_ids")).split(",")
                if item.strip()
            }
            top3_correct += true_template == predicted_template or true_template in alternatives

    reviewed_count = len(reviewed)
    metrics = {
        "reviewed_rows": reviewed_count,
        "description_approval_rate": _ratio(description_approved, reviewed_count),
        "family_accuracy": _ratio(family_correct, reviewed_count),
        "role_accuracy": _ratio(role_correct, reviewed_count),
        "target_valve_precision": _ratio(true_target_and_predicted, predicted_target),
        "critical_field_accuracy": _ratio(critical_correct, critical_total),
        "critical_field_evidence": critical_total,
        "autofill_accuracy": _ratio(autofill_correct, autofill_total),
        "autofill_evidence": autofill_total,
        "template_top1": _ratio(top1_correct, template_total),
        "template_top3": _ratio(top3_correct, template_total),
        "template_evidence": template_total,
    }
    metrics["acceptance"] = {
        "review_complete": reviewed_count >= 600,
        "target_valve_precision": metrics["target_valve_precision"] >= 0.995,
        "critical_field_accuracy": metrics["critical_field_accuracy"] >= 0.99,
        "autofill_accuracy": (
            metrics["autofill_accuracy"] >= 0.99 if autofill_total else False
        ),
        "description_approval_rate": metrics["description_approval_rate"] >= 0.98,
        "template_top1": metrics["template_top1"] >= 0.95,
        "template_top3": metrics["template_top3"] >= 0.99,
    }
    return metrics


def load_review_rows(path: Path | str) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    headers = [_text(value) for value in next(iterator)]
    rows = [
        {header: value for header, value in zip(headers, row)}
        for row in iterator
    ]
    workbook.close()
    return rows


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--gold", type=Path, required=True)
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()
    result = evaluate_review_rows(load_review_rows(args.gold))
    rendered = json.dumps(result, ensure_ascii=False, indent=2)
    if args.output:
        args.output.write_text(rendered, encoding="utf-8")
    print(rendered)


if __name__ == "__main__":
    main()
