"""Evaluate template retrieval on a chronological, leakage-free holdout."""

from __future__ import annotations

import argparse
from collections import Counter, defaultdict
from datetime import date, datetime
import json
from pathlib import Path
import sys
import uuid
from typing import Any, Iterable, Mapping

from openpyxl import load_workbook
from qdrant_client import QdrantClient
from qdrant_client.models import Distance, PointStruct, VectorParams

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from embedder import BGEEmbedder
from service.description_design.retriever import TemplateRetriever
from service.description_design.service import DescriptionDesignService
from service.product_code import normalize_code, parse_product_code
from tools.build_description_assets import (
    BY1_COLUMN,
    CODE_COLUMN,
    DATE_COLUMN,
    DESCRIPTION_COLUMN,
    MATERIAL_COLUMN,
    OUTPUT_COLUMNS,
    SPEC_COLUMN,
)


ATTRIBUTE_FIELDS = OUTPUT_COLUMNS[4:14]
DEFAULT_INPUT = ROOT / "outputs" / "edesc_standardization_v1" / "standardized_orders.xlsx"
DEFAULT_DICTIONARY = ROOT / "data" / "edesc_business_dictionary.json"
DEFAULT_OUTPUT = ROOT / "outputs" / "edesc_standardization_v1" / "template_retrieval_test_metrics.json"


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _date_value(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value)).date()
    except ValueError:
        return None


def _metric_group(results: list[Mapping[str, Any]]) -> dict[str, Any]:
    cases = sum(int(item.get("case_weight", 1)) for item in results)
    recall_hits = sum(
        int(item.get("case_weight", 1))
        for item in results
        if item["target_template_id"] in item.get("candidate_ids", [])[:20]
    )
    accuracy_hits = sum(
        int(item.get("case_weight", 1))
        for item in results
        if item["target_template_id"] in item.get("candidate_ids", [])[:3]
    )
    return {
        "cases": cases,
        "recall_at_20": recall_hits / cases if cases else 0.0,
        "accuracy_at_3": accuracy_hits / cases if cases else 0.0,
        "recall_at_20_hits": recall_hits,
        "accuracy_at_3_hits": accuracy_hits,
    }


def calculate_metrics(results: Iterable[Mapping[str, Any]]) -> dict[str, Any]:
    """Calculate candidate-hit metrics and isolate cold-start templates."""
    rows = list(results)
    seen = [item for item in rows if item.get("seen_in_index")]
    unseen = [item for item in rows if not item.get("seen_in_index")]
    metrics = _metric_group(rows)
    metrics["seen_template"] = _metric_group(seen)
    metrics["unseen_template"] = _metric_group(unseen)
    return metrics


def render_console_result(result: Mapping[str, Any]) -> str:
    """Render results safely for Windows consoles that use a GBK code page."""
    return json.dumps(result, ensure_ascii=True, indent=2)


def _load_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    headers = [_text(value) for value in next(iterator)]
    required = [
        DATE_COLUMN,
        CODE_COLUMN,
        BY1_COLUMN,
        SPEC_COLUMN,
        MATERIAL_COLUMN,
        DESCRIPTION_COLUMN,
        "valve_family",
        "product_role",
        "template_id",
        *ATTRIBUTE_FIELDS,
        "standardized_description",
        "dictionary_version",
    ]
    missing = [column for column in required if column not in headers]
    if missing:
        workbook.close()
        raise ValueError(f"standardized workbook is missing columns: {missing}")
    indexes = {column: headers.index(column) for column in required}
    rows = []
    try:
        for values in iterator:
            role = _text(values[indexes["product_role"]])
            family = _text(values[indexes["valve_family"]])
            template_id = _text(values[indexes["template_id"]])
            description = _text(values[indexes[DESCRIPTION_COLUMN]])
            if role != "valve" or family not in {"butterfly", "check", "gate"}:
                continue
            if not template_id or not description:
                continue
            by1 = normalize_code(values[indexes[BY1_COLUMN]])
            parsed = parse_product_code(
                values[indexes[CODE_COLUMN]],
                by1,
                values[indexes[SPEC_COLUMN]],
                values[indexes[MATERIAL_COLUMN]],
            )
            rows.append(
                {
                    "date": _date_value(values[indexes[DATE_COLUMN]]),
                    "query": description,
                    "by1": by1,
                    "form_code": parsed.form_code,
                    "template_id": template_id,
                    "valve_family": family,
                    "attributes": {
                        field: _text(values[indexes[field]])
                        for field in ATTRIBUTE_FIELDS
                        if _text(values[indexes[field]])
                    },
                    "standardized_description": _text(
                        values[indexes["standardized_description"]]
                    ),
                    "dictionary_version": _text(values[indexes["dictionary_version"]]),
                }
            )
    finally:
        workbook.close()
    return rows


def _build_templates(rows: Iterable[Mapping[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, list[Mapping[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[str(row["template_id"])].append(row)
    templates = []
    for template_id, members in grouped.items():
        attributes = dict(members[0]["attributes"])
        descriptions = Counter(
            _text(item["standardized_description"])
            for item in members
            if _text(item["standardized_description"])
        )
        dates = [item["date"] for item in members if item["date"]]
        family = str(members[0]["valve_family"])
        semantic_attributes = " ".join(
            f"{field} {value}" for field, value in sorted(attributes.items()) if field != "size"
        )
        templates.append(
            {
                "template_id": template_id,
                "valve_family": family,
                "product_role": "valve",
                "attributes": attributes,
                "standardized_description": descriptions.most_common(1)[0][0]
                if descriptions
                else "",
                "supported_by1": sorted({str(item["by1"]) for item in members if item["by1"]}),
                "form_codes": sorted(
                    {str(item["form_code"]) for item in members if item["form_code"]}
                ),
                "support": len(members),
                "date_range": {
                    "start": min(dates).isoformat() if dates else "",
                    "end": max(dates).isoformat() if dates else "",
                },
                "dictionary_version": str(members[0]["dictionary_version"]),
                "semantic_text": f"{family} {semantic_attributes}".strip(),
            }
        )
    return templates


def _build_in_memory_retriever(
    templates: list[Mapping[str, Any]], embedder: Any
) -> TemplateRetriever:
    if not templates:
        raise ValueError("no templates available before the training cutoff")
    vectors = embedder.encode([str(item["semantic_text"]) for item in templates], batch_size=64)
    if not vectors or not vectors[0]:
        raise RuntimeError("embedding service returned no template vectors")
    client = QdrantClient(":memory:")
    collection = "template_retrieval_evaluation"
    client.create_collection(
        collection_name=collection,
        vectors_config=VectorParams(size=len(vectors[0]), distance=Distance.COSINE),
    )
    points = [
        PointStruct(
            id=str(uuid.uuid5(uuid.NAMESPACE_URL, f"template-evaluation:{item['template_id']}")),
            vector=vector,
            payload=dict(item),
        )
        for item, vector in zip(templates, vectors)
    ]
    for start in range(0, len(points), 256):
        client.upsert(collection, points=points[start : start + 256], wait=True)
    return TemplateRetriever(client, embedder, collection, candidate_limit=100)


def _deduplicate(rows: Iterable[Mapping[str, Any]]) -> list[Mapping[str, Any]]:
    selected: dict[tuple[str, str, str, str], dict[str, Any]] = {}
    for row in rows:
        key = (row["query"], row["by1"], row["form_code"], row["template_id"])
        if key not in selected:
            selected[key] = {**row, "case_weight": 0}
        selected[key]["case_weight"] += 1
    return list(selected.values())


def _evaluate_cases(
    rows: list[Mapping[str, Any]],
    retriever: TemplateRetriever,
    service: DescriptionDesignService,
    indexed_template_ids: set[str],
) -> list[dict[str, Any]]:
    designs = [
        service.engine.design(
            str(row["query"]),
            by1="",
            form_code=str(row["form_code"]),
            inferred={},
        )
        for row in rows
    ]
    vectors = retriever.embedder.encode([str(row["query"]) for row in rows], batch_size=64)
    results = []
    for row, design, vector in zip(rows, designs, vectors):
        candidates = retriever.retrieve_with_vector(
            vector,
            str(row["query"]),
            design,
            by1="",
            form_code=str(row["form_code"]),
            top_k=20,
        )
        results.append(
            {
                "target_template_id": str(row["template_id"]),
                "candidate_ids": [item["template_id"] for item in candidates],
                "seen_in_index": str(row["template_id"]) in indexed_template_ids,
                "case_weight": int(row.get("case_weight", 1)),
                "input": {
                    "query": str(row["query"]),
                    "form_code": str(row["form_code"]),
                },
                "expected_by1": str(row["by1"]),
                "candidates": [
                    {
                        "template_id": item["template_id"],
                        "by1": list(item.get("by1", [])),
                        "form_codes": list(item.get("form_codes", [])),
                    }
                    for item in candidates[:3]
                ],
            }
        )
    return results


def evaluate_template_retrieval(
    input_path: Path,
    dictionary_path: Path,
    *,
    train_until: date,
    test_from: date,
) -> dict[str, Any]:
    """Run row-level and de-duplicated template-retrieval evaluation."""
    all_rows = _load_rows(input_path)
    train_rows = [row for row in all_rows if row["date"] and row["date"] <= train_until]
    test_rows = [row for row in all_rows if row["date"] and row["date"] >= test_from]
    templates = _build_templates(train_rows)
    embedder = BGEEmbedder(timeout=120)
    retriever = _build_in_memory_retriever(templates, embedder)
    service = DescriptionDesignService.from_dictionary_path(dictionary_path)
    indexed_template_ids = {str(template["template_id"]) for template in templates}
    deduplicated_rows = _deduplicate(test_rows)
    row_results = _evaluate_cases(
        deduplicated_rows, retriever, service, indexed_template_ids
    )
    deduplicated_results = [{**item, "case_weight": 1} for item in row_results]
    return {
        "evaluation": {
            "train_until": train_until.isoformat(),
            "test_from": test_from.isoformat(),
            "retrieval_mode": "bge_structured_without_reranker",
            "input_fields": ["query", "form_code"],
            "candidate_output_fields": ["template_id", "by1", "form_codes"],
            "training_template_count": len(templates),
            "test_row_count": len(test_rows),
            "test_deduplicated_case_count": len(deduplicated_results),
        },
        "row_level": calculate_metrics(row_results),
        "deduplicated": calculate_metrics(deduplicated_results),
        "prediction_examples": row_results[:20],
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--dictionary", type=Path, default=DEFAULT_DICTIONARY)
    parser.add_argument("--train-until", default="2024-06-30")
    parser.add_argument("--test-from", default="2024-07-01")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()
    result = evaluate_template_retrieval(
        args.input,
        args.dictionary,
        train_until=date.fromisoformat(args.train_until),
        test_from=date.fromisoformat(args.test_from),
    )
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(render_console_result(result))


if __name__ == "__main__":
    main()
