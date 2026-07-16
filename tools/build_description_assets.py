"""Build valve-description dictionaries, normalized orders, templates, and gold data."""

from __future__ import annotations

import argparse
from collections import Counter, defaultdict
from datetime import date, datetime
import hashlib
import json
from pathlib import Path
import random
import sys
from typing import Any, Iterable, Mapping

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from service.description_design import BusinessDictionary, DescriptionDesignEngine
from service.product_code import parse_product_code


DATE_COLUMN = "合同日期"
CODE_COLUMN = "产品编码"
BY1_COLUMN = "品种"
SPEC_COLUMN = "规格"
MATERIAL_COLUMN = "材质分类"
DESCRIPTION_COLUMN = "英文描述"

REQUIRED_COLUMNS = [
    DATE_COLUMN,
    CODE_COLUMN,
    BY1_COLUMN,
    SPEC_COLUMN,
    MATERIAL_COLUMN,
    DESCRIPTION_COLUMN,
]
REQUIRED_DICTIONARY_SHEETS = [
    "词语映射",
    "标准字段值",
    "货描模板",
    "品种形式规则",
    "例外规则",
    "待确认项",
]

OUTPUT_COLUMNS = [
    "row_id",
    "valve_family",
    "product_role",
    "template_id",
    "body_material",
    "connection",
    "structure",
    "seat_material",
    "closure_material",
    "actuation",
    "pressure",
    "standard_certification",
    "size",
    "special_requirements",
    "standardized_description",
    "inferred_fields",
    "confidence",
    "confidence_score",
    "warnings",
    "dictionary_version",
]

CRITICAL_FIELDS = {
    "body_material",
    "seat_material",
    "closure_material",
    "pressure",
    "standard_certification",
    "size",
}
TEMPLATE_DEFINITIONS = {
    "butterfly": {
        "template_id": "BUTTERFLY-BASE",
        "field_order": "body_material,connection,structure,valve_family,seat_material,closure_material,actuation,special_requirements,pressure,standard_certification,size",
        "required_fields": "body_material,connection,size",
    },
    "check": {
        "template_id": "CHECK-BASE",
        "field_order": "body_material,connection,structure,valve_family,seat_material,closure_material,actuation,special_requirements,pressure,standard_certification,size",
        "required_fields": "body_material,connection,size",
    },
    "gate": {
        "template_id": "GATE-BASE",
        "field_order": "body_material,connection,structure,valve_family,seat_material,closure_material,actuation,special_requirements,pressure,standard_certification,size",
        "required_fields": "body_material,connection,size",
    },
}


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _date_value(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value)).date()
    except ValueError:
        return None


def derive_mature_rules(
    records: Iterable[Mapping[str, Any]],
) -> dict[tuple[str, str], dict[str, dict[str, Any]]]:
    """Derive only temporally stable by1 and by1/form attribute rules."""
    counters: dict[
        tuple[str, str],
        dict[str, dict[str, Counter[str]]],
    ] = defaultdict(lambda: defaultdict(lambda: {"train": Counter(), "validation": Counter()}))
    for record in records:
        record_date = _date_value(record.get("date"))
        if record_date is None:
            continue
        if record_date <= date(2023, 12, 31):
            window = "train"
        elif record_date <= date(2024, 6, 30):
            window = "validation"
        else:
            continue
        by1 = _text(record.get("by1")).upper()
        form_code = _text(record.get("form_code")).upper()
        if not by1:
            continue
        for key in [(by1, ""), (by1, form_code)] if form_code else [(by1, "")]:
            for field, item in record.get("attributes", {}).items():
                value = _text(item.get("value") if isinstance(item, Mapping) else item)
                if value:
                    counters[key][field][window][value] += 1

    rules: dict[tuple[str, str], dict[str, dict[str, Any]]] = {}
    for key, field_windows in counters.items():
        accepted = {}
        for field, windows in field_windows.items():
            train = windows["train"]
            validation = windows["validation"]
            train_total = sum(train.values())
            validation_total = sum(validation.values())
            if train_total < 20 or validation_total < 5:
                continue
            train_value, train_support = train.most_common(1)[0]
            validation_value, validation_support = validation.most_common(1)[0]
            train_purity = train_support / train_total
            validation_purity = validation_support / validation_total
            required_train_purity = 1.0 if field in CRITICAL_FIELDS else 0.99
            if train_value != validation_value:
                continue
            if train_purity < required_train_purity or validation_purity != 1.0:
                continue
            accepted[field] = {
                "value": train_value,
                "source": "mature_rule",
                "confidence": 0.99,
                "train_support": train_total,
                "validation_support": validation_total,
                "train_purity": round(train_purity, 6),
                "validation_purity": round(validation_purity, 6),
            }
        if accepted:
            rules[key] = accepted
    return rules


def _read_rows(source: Path):
    workbook = load_workbook(source, read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    headers = [_text(value) for value in next(iterator)]
    missing = [column for column in REQUIRED_COLUMNS if column not in headers]
    if missing:
        workbook.close()
        raise ValueError(f"missing required columns: {', '.join(missing)}")
    indexes = {column: headers.index(column) for column in REQUIRED_COLUMNS}
    return workbook, headers, indexes, iterator


def _first_pass(source: Path, engine: DescriptionDesignEngine) -> list[dict[str, Any]]:
    workbook, _, indexes, rows = _read_rows(source)
    records = []
    try:
        for values in rows:
            description = _text(values[indexes[DESCRIPTION_COLUMN]])
            by1 = _text(values[indexes[BY1_COLUMN]])
            spec = _text(values[indexes[SPEC_COLUMN]])
            material = _text(values[indexes[MATERIAL_COLUMN]])
            parsed = parse_product_code(
                values[indexes[CODE_COLUMN]], by1, spec, material
            )
            result = engine.design(description, by1=by1, form_code=parsed.form_code)
            if result["product_role"] != "valve" or not result["valve_family"]:
                continue
            records.append(
                {
                    "date": values[indexes[DATE_COLUMN]],
                    "by1": by1.upper(),
                    "form_code": parsed.form_code,
                    "attributes": result["attributes"],
                }
            )
    finally:
        workbook.close()
    return records


def _rule_attributes(
    rules: Mapping[tuple[str, str], Mapping[str, Mapping[str, Any]]],
    by1: str,
    form_code: str,
) -> dict[str, Mapping[str, Any]]:
    result = dict(rules.get((by1.upper(), ""), {}))
    result.update(rules.get((by1.upper(), form_code.upper()), {}))
    return result


def _attribute_value(result: Mapping[str, Any], field: str) -> str | None:
    item = result.get("attributes", {}).get(field)
    return item.get("value") if item else None


def _template_payload(
    aggregate: Mapping[str, Any],
    dictionary_version: str,
) -> dict[str, Any]:
    attributes = dict(aggregate["attributes"])
    attributes.pop("size", None)
    return {
        "template_id": aggregate["template_id"],
        "valve_family": aggregate["valve_family"],
        "product_role": "valve",
        "attributes": attributes,
        "standardized_description": aggregate["example_description"],
        "supported_by1": sorted(aggregate["by1"]),
        "form_codes": sorted(aggregate["form_codes"]),
        "support": aggregate["support"],
        "date_min": aggregate["date_min"],
        "date_max": aggregate["date_max"],
        "examples": aggregate["examples"][:3],
        "dictionary_version": dictionary_version,
        "semantic_text": " | ".join(
            [
                aggregate["valve_family"],
                *[f"{key}: {value}" for key, value in sorted(attributes.items())],
            ]
        ),
    }


def _update_template(
    templates: dict[str, dict[str, Any]],
    result: Mapping[str, Any],
    *,
    by1: str,
    form_code: str,
    description: str,
    row_date: object,
) -> None:
    template_id = result.get("template_id")
    if not template_id or result.get("product_role") != "valve":
        return
    attributes = {
        field: item["value"]
        for field, item in result.get("attributes", {}).items()
    }
    aggregate = templates.setdefault(
        template_id,
        {
            "template_id": template_id,
            "valve_family": result["valve_family"],
            "attributes": attributes,
            "example_description": result["standardized_description"],
            "by1": set(),
            "form_codes": set(),
            "support": 0,
            "date_min": None,
            "date_max": None,
            "examples": [],
        },
    )
    aggregate["support"] += 1
    if by1:
        aggregate["by1"].add(by1.upper())
    if form_code:
        aggregate["form_codes"].add(form_code.upper())
    if description and len(aggregate["examples"]) < 3:
        aggregate["examples"].append(description)
    parsed_date = _date_value(row_date)
    if parsed_date:
        date_text = parsed_date.isoformat()
        aggregate["date_min"] = min(aggregate["date_min"], date_text) if aggregate["date_min"] else date_text
        aggregate["date_max"] = max(aggregate["date_max"], date_text) if aggregate["date_max"] else date_text


def _reservoir_add(
    reservoirs: dict[str, list[dict[str, Any]]],
    key: str,
    item: dict[str, Any],
    row_id: int,
    limit: int = 500,
) -> None:
    bucket = reservoirs.setdefault(key, [])
    if len(bucket) < limit:
        bucket.append(item)
        return
    rng = random.Random(row_id * 104729)
    replacement = rng.randrange(row_id)
    if replacement < limit:
        bucket[replacement] = item


def _write_dictionary_workbook(
    path: Path,
    dictionary: BusinessDictionary,
    mature_rules: Mapping[tuple[str, str], Mapping[str, Mapping[str, Any]]],
    templates: Iterable[Mapping[str, Any]],
    pending: Counter[str],
    source_path: Path | None = None,
) -> None:
    if source_path and source_path.exists():
        workbook = load_workbook(source_path)
        for sheet_name in ["品种形式规则", "待确认项"]:
            if sheet_name in workbook.sheetnames:
                workbook.remove(workbook[sheet_name])
        rule_sheet = workbook.create_sheet("品种形式规则")
        rule_sheet.append(["by1", "form_code", "field", "value", "confidence", "train_support", "validation_support", "train_purity", "validation_purity", "active"])
        for (by1, form_code), fields in sorted(mature_rules.items()):
            for field, item in sorted(fields.items()):
                rule_sheet.append([
                    by1, form_code, field, item["value"], item["confidence"],
                    item["train_support"], item["validation_support"],
                    item["train_purity"], item["validation_purity"], True,
                ])
        pending_sheet = workbook.create_sheet("待确认项")
        pending_sheet.append(["description", "support", "review_status", "notes"])
        for description, support in pending.most_common(500):
            pending_sheet.append([description, support, "pending", ""])
        for sheet in [rule_sheet, pending_sheet]:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E78")
        workbook.save(path)
        return

    workbook = Workbook()
    workbook.remove(workbook.active)

    terms = workbook.create_sheet("词语映射")
    terms.append(["raw_term", "canonical_term", "field", "family_scope", "priority", "active", "notes"])
    for rule in dictionary.rules:
        terms.append([rule.term, rule.value, rule.field, ",".join(rule.families), rule.priority, True, ""])

    values = workbook.create_sheet("标准字段值")
    values.append(["field", "canonical_value", "family_scope", "active", "notes"])
    distinct_values = sorted({(rule.field, rule.value, ",".join(rule.families)) for rule in dictionary.rules})
    for field, value, scope in distinct_values:
        values.append([field, value, scope, True, ""])

    template_sheet = workbook.create_sheet("货描模板")
    template_sheet.append(["template_id", "valve_family", "field_order", "required_fields", "active", "observed_support"])
    supports = Counter(item["valve_family"] for item in templates)
    for family, item in TEMPLATE_DEFINITIONS.items():
        template_sheet.append([item["template_id"], family, item["field_order"], item["required_fields"], True, supports[family]])

    rule_sheet = workbook.create_sheet("品种形式规则")
    rule_sheet.append(["by1", "form_code", "field", "value", "confidence", "train_support", "validation_support", "train_purity", "validation_purity", "active"])
    for (by1, form_code), fields in sorted(mature_rules.items()):
        for field, item in sorted(fields.items()):
            rule_sheet.append([
                by1, form_code, field, item["value"], item["confidence"],
                item["train_support"], item["validation_support"],
                item["train_purity"], item["validation_purity"], True,
            ])

    exceptions = workbook.create_sheet("例外规则")
    exceptions.append(["match_field", "match_value", "override_field", "override_value", "reason", "active"])

    pending_sheet = workbook.create_sheet("待确认项")
    pending_sheet.append(["description", "support", "review_status", "notes"])
    for description, support in pending.most_common(500):
        pending_sheet.append([description, support, "pending", ""])

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
    workbook.save(path)


def compile_dictionary_workbook(
    workbook_path: Path,
    output_path: Path,
    version: str,
) -> dict[str, Any]:
    """Compile the reviewable Excel source into deterministic runtime JSON."""
    source_hash = hashlib.sha256(workbook_path.read_bytes()).hexdigest()
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    payload: dict[str, Any] = {
        "version": version,
        "source_file": workbook_path.name,
        "source_sha256": source_hash,
        "sheets": {},
    }
    for sheet_name in REQUIRED_DICTIONARY_SHEETS:
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        headers = [_text(value) for value in next(rows)]
        payload["sheets"][sheet_name] = [
            {header: value for header, value in zip(headers, row)}
            for row in rows
            if any(value is not None for value in row)
        ]
    workbook.close()
    output_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )
    return payload


def _select_gold_rows(reservoirs: Mapping[str, list[dict[str, Any]]]) -> list[dict[str, Any]]:
    selected = []
    for family in ["butterfly", "check", "gate"]:
        rows = list(reservoirs.get(family, []))
        rows.sort(key=lambda item: (item["status"] != "partial", item["row_id"]))
        selected.extend(rows[:150])
    excluded = list(reservoirs.get("accessory", [])) + list(reservoirs.get("spare_part", []))
    excluded.sort(key=lambda item: item["row_id"])
    selected.extend(excluded[:75])
    other = sorted(reservoirs.get("other", []), key=lambda item: item["row_id"])
    selected.extend(other[:75])
    return selected


def _write_gold_workbook(path: Path, rows: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "人工金标"
    attribute_fields = [
        "body_material", "connection", "structure", "seat_material",
        "closure_material", "actuation", "pressure", "standard_certification",
        "size", "special_requirements",
    ]
    headers = [
        "row_id", "description", "by1", "form_code", "predicted_family",
        "predicted_role", "template_id", "alternative_template_ids",
        *attribute_fields, "inferred_fields", "predicted_description",
        "confidence", "warnings", "approved", "corrected_family",
        "corrected_role", "corrected_template_id",
        *[f"corrected_{field}" for field in attribute_fields],
        "corrected_description", "reviewer_notes",
    ]
    sheet.append(headers)
    for item in rows:
        sheet.append([
            item.get("row_id"), item.get("description"), item.get("by1"),
            item.get("form_code"), item.get("valve_family"),
            item.get("product_role"), item.get("template_id"),
            ",".join(
                str(candidate.get("template_id"))
                for candidate in item.get("alternatives", [])
                if candidate.get("template_id")
            ),
            *[
                (item.get("attributes", {}).get(field) or {}).get("value")
                for field in attribute_fields
            ],
            ",".join(item.get("inferred_fields", [])),
            item.get("standardized_description"), item.get("confidence"),
            item.get("warnings"), "", "", "", "",
            *["" for _ in attribute_fields], "", "",
        ])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="548235")
    workbook.save(path)


def build_assets(
    source: Path | str,
    output_dir: Path | str,
    *,
    dictionary_version: str,
    dictionary_source: Path | str | None = None,
) -> dict[str, Any]:
    """Build every offline asset from one immutable order workbook."""
    source = Path(source)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    standardized_path = output_dir / "standardized_orders.xlsx"
    dictionary_path = output_dir / "edesc_business_dictionary.xlsx"
    dictionary_json_path = output_dir / "edesc_business_dictionary.json"
    template_index_path = output_dir / "edesc_template_index.json"
    gold_path = output_dir / "edesc_gold_review.xlsx"
    report_path = output_dir / "edesc_build_report.json"
    effective_dictionary_source = Path(dictionary_source) if dictionary_source else None
    if effective_dictionary_source is None and dictionary_path.exists():
        effective_dictionary_source = dictionary_path
    dictionary = (
        BusinessDictionary.from_excel(effective_dictionary_source)
        if effective_dictionary_source and effective_dictionary_source.exists()
        else BusinessDictionary.default()
    )
    engine = DescriptionDesignEngine(dictionary)
    mature_rules = derive_mature_rules(_first_pass(source, engine))

    source_workbook, source_headers, indexes, source_rows = _read_rows(source)
    output_workbook = Workbook(write_only=True)
    output_sheet = output_workbook.create_sheet("standardized_orders")
    output_sheet.append(source_headers + OUTPUT_COLUMNS)
    templates: dict[str, dict[str, Any]] = {}
    reservoirs: dict[str, list[dict[str, Any]]] = {}
    pending: Counter[str] = Counter()
    counts = Counter()
    try:
        for row_id, values in enumerate(source_rows, start=1):
            description = _text(values[indexes[DESCRIPTION_COLUMN]])
            by1 = _text(values[indexes[BY1_COLUMN]])
            spec = _text(values[indexes[SPEC_COLUMN]])
            material = _text(values[indexes[MATERIAL_COLUMN]])
            parsed = parse_product_code(values[indexes[CODE_COLUMN]], by1, spec, material)
            inferred = _rule_attributes(mature_rules, by1, parsed.form_code)
            result = engine.design(
                description,
                by1=by1,
                form_code=parsed.form_code,
                inferred=inferred,
            )
            warnings = list(result.get("warnings", []))
            if parsed.status != "ok":
                warnings.append(f"product_code_{parsed.status}")
            counts["rows"] += 1
            counts[f"family_{result.get('valve_family') or 'other'}"] += 1
            counts[f"role_{result['product_role']}"] += 1
            counts[f"status_{result['status']}"] += 1
            if result["status"] == "partial" and description:
                pending[description] += 1

            extras = [
                row_id,
                result.get("valve_family"),
                result["product_role"],
                result.get("template_id"),
                *[_attribute_value(result, field) for field in OUTPUT_COLUMNS[4:14]],
                result.get("standardized_description"),
                ",".join(result.get("inferred_fields", [])),
                result.get("confidence"),
                result.get("confidence_score"),
                ";".join(warnings),
                dictionary_version,
            ]
            output_sheet.append(list(values) + extras)
            _update_template(
                templates,
                result,
                by1=by1,
                form_code=parsed.form_code,
                description=description,
                row_date=values[indexes[DATE_COLUMN]],
            )
            reservoir_key = result.get("valve_family") or result["product_role"]
            review_item = {
                "row_id": row_id,
                "description": description,
                "by1": by1,
                "form_code": parsed.form_code,
                **result,
                "warnings": ";".join(warnings),
            }
            _reservoir_add(reservoirs, reservoir_key, review_item, row_id)
    finally:
        source_workbook.close()
    output_workbook.save(standardized_path)

    template_payloads = [
        _template_payload(item, dictionary_version)
        for item in sorted(templates.values(), key=lambda value: value["template_id"])
    ]
    template_index_path.write_text(
        json.dumps(
            {
                "version": dictionary_version,
                "source": source.name,
                "templates": template_payloads,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    _write_dictionary_workbook(
        dictionary_path,
        engine.dictionary,
        mature_rules,
        template_payloads,
        pending,
        source_path=effective_dictionary_source,
    )
    compile_dictionary_workbook(dictionary_path, dictionary_json_path, dictionary_version)
    gold_rows = _select_gold_rows(reservoirs)
    _write_gold_workbook(gold_path, gold_rows)

    report = {
        "source": source.name,
        "dictionary_version": dictionary_version,
        "counts": dict(counts),
        "mature_rule_groups": len(mature_rules),
        "templates": len(template_payloads),
        "gold_rows": len(gold_rows),
        "acceptance": {
            "status": "pending_human_review",
            "valve_precision_target": 0.995,
            "critical_field_accuracy_target": 0.99,
            "autofill_accuracy_target": 0.99,
            "description_approval_target": 0.98,
            "template_top1_target": 0.95,
            "template_top3_target": 0.99,
        },
    }
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return {
        "standardized_workbook": standardized_path,
        "dictionary_workbook": dictionary_path,
        "dictionary_json": dictionary_json_path,
        "template_index": template_index_path,
        "gold_workbook": gold_path,
        "report": report_path,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--dictionary-version", default=datetime.now().strftime("%Y.%m.%d"))
    parser.add_argument("--dictionary-source", type=Path)
    args = parser.parse_args()
    result = build_assets(
        args.input,
        args.output_dir,
        dictionary_version=args.dictionary_version,
        dictionary_source=args.dictionary_source,
    )
    print(json.dumps({key: str(value) for key, value in result.items()}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
